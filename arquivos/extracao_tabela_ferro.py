import math
import os
from datetime import datetime
from TQS import TQSDwg, TQSGeo, TQSEag, TQSJan
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TOLERANCIA = 10
NIVEL_VIGA = 228
NIVEL_PILAR = 227


def mapear_contorno(dwg):
    segmentos = []

    dwg.iterator.Begin()
    while True:
        itipo = dwg.iterator.Next()

        if itipo == TQSDwg.DWGTYPE_EOF:
            break
        if itipo not in (TQSDwg.DWGTYPE_LINE, TQSDwg.DWGTYPE_POLYLINE):
            continue
        if dwg.iterator.level not in (NIVEL_VIGA, NIVEL_PILAR):
            continue

        if itipo == TQSDwg.DWGTYPE_LINE:
            segmentos.append((dwg.iterator.x1, dwg.iterator.y1, dwg.iterator.x2, dwg.iterator.y2))
            continue

        npts = dwg.iterator.xySize
        if npts < 2:
            continue

        x_ant, y_ant = dwg.iterator.GetPolylinePt(0)
        for ipt in range(1, npts):
            x_cur, y_cur = dwg.iterator.GetPolylinePt(ipt)
            segmentos.append((x_ant, y_ant, x_cur, y_cur))
            x_ant, y_ant = x_cur, y_cur

    return segmentos


def extrair_ferros(dwg, eag=None):
    ferros = []

    dwg.iterator.Begin()
    while True:
        itipo = dwg.iterator.Next()

        if itipo == TQSDwg.DWGTYPE_EOF:
            break
        if itipo != TQSDwg.DWGTYPE_OBJECT:
            continue
        if dwg.iterator.objectName != "IPOFER":
            continue

        rebar = dwg.iterator.smartRebar
        ferro_tipo = rebar.type
        info_descr = []

        numdescr = rebar.RebarScheduleNumDescr()
        for idescr in range(numdescr):
            (
                ipos,
                bitola,
                nfer,
                mult,
                itpcorbar,
                ivar,
                rdval,
                rddsc,
                compr,
                igane,
                igand,
                observ,
                ilance,
                itipo99,
                icftppata,
                icorrido,
                iluvai,
                iluvaf,
            ) = rebar.RebarScheduleInfo(idescr)
            info_descr.append({"ipos": ipos, "nfer": nfer})

        pontos_insercao = []
        if ferro_tipo == TQSDwg.ICPFGN:
            n_pontos = rebar.GetGenRebarPoints()
            (
                xins,
                yins,
                angins,
                escxy,
                identfer,
                identdobr,
                ipatas,
                iexplodir,
                inivel,
                iestilo,
                icor,
            ) = rebar.GetInsertionData(0)

            for ipt in range(n_pontos):
                if ipt not in (0, 1, n_pontos - 2, n_pontos - 1):
                    continue

                xpt_local, ypt_local, zpt, iarco, identdobr, indfrt = rebar.GetGenRebarPoint(ipt)
                x_rot, y_rot = TQSGeo.Rotate(xpt_local, ypt_local, angins, 0.0, 0.0)
                pontos_insercao.append(((x_rot * escxy) + xins, (y_rot * escxy) + yins))
        else:
            n_ins = rebar.GetInsertionNumber()
            for idx_ins in range(n_ins):
                resultado = rebar.GetInsertionPoints(idx_ins)
                npts = resultado[0] if isinstance(resultado, tuple) else resultado
                for ipt in range(npts):
                    pt = rebar.GetInsertionPoint(idx_ins, ipt)
                    pontos_insercao.append((pt[0], pt[1]))

        alternado = rebar.alternatingMode == TQSDwg.ICPCAL

        ferros.append(
            {
                "tipo": ferro_tipo,
                "bitola_mm": rebar.diameter,
                "info_descr": info_descr,
                "pontos": pontos_insercao,
                "espacamento": rebar.spacing,
                "alternado": alternado,
            }
        )

    return ferros


def distancia_ponto_segmento(xp, yp, x1, y1, x2, y2):
    dx = x2 - x1
    dy = y2 - y1
    comp2 = dx * dx + dy * dy

    if comp2 == 0:
        return math.sqrt((xp - x1) ** 2 + (yp - y1) ** 2)

    t = ((xp - x1) * dx + (yp - y1) * dy) / comp2
    t = max(0.0, min(1.0, t))
    px = x1 + t * dx
    py = y1 + t * dy
    return math.sqrt((xp - px) ** 2 + (yp - py) ** 2)


def ponto_na_borda(xp, yp, segmentos):
    limite_superior = TOLERANCIA

    for x1l, y1l, x2l, y2l in segmentos:
        if distancia_ponto_segmento(xp, yp, x1l, y1l, x2l, y2l) <= limite_superior:
            return True

    return False


def classificar_ferros(ferros, segmentos):
    for ferro in ferros:
        ferro["eh_borda"] = any(
            ponto_na_borda(xp, yp, segmentos) for xp, yp in ferro["pontos"]
        )


def gerar_relatorio(ferros, arquivo_excel):
    ferros_ordenados = sorted(ferros, key=lambda f: f["info_descr"][0]["ipos"] if f["info_descr"] else 999)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ferros"

    cinza = PatternFill(patternType="solid", fgColor="FFD9D9D9", bgColor="FFFFFFFF")
    azul_claro = PatternFill(patternType="solid", fgColor="FFBDD7EE", bgColor="FFFFFFFF")
    verde = PatternFill(patternType="solid", fgColor="FFC6EFCE", bgColor="FFFFFFFF")
    amarelo = PatternFill(patternType="solid", fgColor="FFFFEB9C", bgColor="FFFFFFFF")
    vermelho = PatternFill(patternType="solid", fgColor="FFFF9999", bgColor="FFFFFFFF")
    centro = Alignment(horizontal="center", vertical="center")
    borda_fina = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def estilizar(cell, fill=None, bold=False, align=True):
        if fill:
            cell.fill = fill
        if bold:
            cell.font = Font(bold=True)
        if align:
            cell.alignment = centro
        cell.border = borda_fina

    cabecalho = ["Posição", "Quantidade", "Espaçamento (cm)", "Compr. Faixa (cm)", "Tipo"]
    ws.append(cabecalho)
    for col, _ in enumerate(cabecalho, start=1):
        estilizar(ws.cell(row=1, column=col), fill=cinza, bold=True)

    cores_tipo = {"BORDA": vermelho, "ALTERNADO": amarelo, "INTERNO": azul_claro}

    soma_geral = 0.0
    soma_alt = 0.0
    soma_simp = 0.0
    soma_borda = 0.0
    pos_anterior = None

    for ferro in ferros_ordenados[0:-2]:
        if ferro["eh_borda"]:
            tipo = "BORDA"
        elif ferro["alternado"]:
            tipo = "ALTERNADO"
        else:
            tipo = "INTERNO"

        nfer = ferro["info_descr"][0]["nfer"] if ferro["info_descr"] else 0
        espacamento = ferro["espacamento"]
        comp_faixa_total = nfer * espacamento

        ipos = ferro["info_descr"][0]["ipos"] if ferro["info_descr"] else None
        if pos_anterior is not None and ipos != pos_anterior:
            ws.append(["", "", "", "", ""])
        pos_anterior = ipos

        for info in ferro["info_descr"]:
            ws.append([f"N{info['ipos']}", info["nfer"], int(round(espacamento)), int(round(comp_faixa_total)), tipo])
            fill = cores_tipo["BORDA" if ferro["eh_borda"] else ("ALTERNADO" if ferro["alternado"] else "INTERNO")]
            for col in range(1, 6):
                estilizar(ws.cell(ws.max_row, col), fill=fill)

        soma_geral += comp_faixa_total
        if ferro["eh_borda"]:
            soma_borda += comp_faixa_total
        elif ferro["alternado"]:
            soma_alt += comp_faixa_total
        else:
            soma_simp += comp_faixa_total

    ws.append(["", "", "", "", ""])
    totais = [
        ("Total Geral (cm)", soma_geral, verde),
        ("Total Alternados (cm)", soma_alt, amarelo),
        ("Total Simples (cm)", soma_simp, azul_claro),
        ("Total Borda (cm)", soma_borda, vermelho),
    ]
    for label, valor, fill in totais:
        ws.append([label, "", "", int(round(valor)), ""])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        for col in range(1, 6):
            estilizar(ws.cell(r, col), fill=fill, bold=True)

    # Bloco auxiliar à direita da tabela: imagens e resultados DETS/DETAL.
    primeira_coluna_imagem = 9  # duas colunas de espaçamento após a tabela (A:E)
    coluna_imagem = get_column_letter(primeira_coluna_imagem)
    primeira_linha_imagem = 12
    base_dir = os.path.dirname(__file__)
    candidatos_imgs = [
        os.path.join(base_dir, "imgs"),
    ]
    imagens_dir = next((p for p in candidatos_imgs if os.path.isdir(p)), candidatos_imgs[-1])
    caminho_det_s = os.path.join(imagens_dir, "detS.png")
    caminho_det_al = os.path.join(imagens_dir, "detAL.png")

    valor_det_s = soma_simp + (soma_borda / 2.0)
    valor_det_al = soma_alt

    if os.path.exists(caminho_det_s):
        img_det_s = Image(caminho_det_s)
        img_det_s.width *= 0.5
        img_det_s.height *= 0.5
        ws.add_image(img_det_s, f"{coluna_imagem}{primeira_linha_imagem}")
        linhas_ocupadas_det_s = max(12, int(img_det_s.height / 20))
    else:
        linhas_ocupadas_det_s = 14
        ws.cell(row=primeira_linha_imagem, column=primeira_coluna_imagem, value="Imagem detS não encontrada")

    linha_valor_det_s = primeira_linha_imagem + linhas_ocupadas_det_s - 3
    ws.cell(row=linha_valor_det_s, column=primeira_coluna_imagem, value="Total Simples + Borda")
    ws.cell(row=linha_valor_det_s + 1, column=primeira_coluna_imagem, value=int(round(valor_det_s)))

    segunda_linha_imagem = linha_valor_det_s + 4
    if os.path.exists(caminho_det_al):
        img_det_al = Image(caminho_det_al)
        img_det_al.width *= 0.5
        img_det_al.height *= 0.5
        ws.add_image(img_det_al, f"{coluna_imagem}{segunda_linha_imagem}")
        linhas_ocupadas_det_al = max(12, int(img_det_al.height / 20))
    else:
        linhas_ocupadas_det_al = 14
        ws.cell(row=segunda_linha_imagem, column=primeira_coluna_imagem, value="Imagem detAL não encontrada")

    linha_valor_det_al = segunda_linha_imagem + linhas_ocupadas_det_al - 2
    ws.cell(row=linha_valor_det_al, column=primeira_coluna_imagem, value="Total Alternados")
    ws.cell(row=linha_valor_det_al + 1, column=primeira_coluna_imagem, value=int(round(valor_det_al)))

    for linha in (linha_valor_det_s, linha_valor_det_s + 1, linha_valor_det_al, linha_valor_det_al + 1):
        estilizar(ws.cell(linha, primeira_coluna_imagem), fill=cinza if linha in (linha_valor_det_s, linha_valor_det_al) else None, bold=True)

    larguras = [14, 13, 18, 20, 14]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    largura_coluna_imagem = 24
    if os.path.exists(caminho_det_al):
        largura_coluna_imagem = img_det_al.width / 7
    ws.column_dimensions[coluna_imagem].width = largura_coluna_imagem

    os.makedirs(os.path.dirname(arquivo_excel), exist_ok=True)
    try:
        wb.save(arquivo_excel)
        return arquivo_excel
    except PermissionError:
        return None


def rodar_script(eag=None, tqsjan=None):
    if eag is None:
        eag = TQSEag.TQSEag()
    if tqsjan is None:
        tqsjan = TQSJan.TQSJan()

    dwg = tqsjan.dwg
    dwg_path = dwg.file.Name()
    draw_name = os.path.basename(dwg_path)
    dwg_dir = os.path.dirname(dwg_path)
    arquivo_excel = os.path.join(dwg_dir, f"Ferro Corrido - {draw_name}.xlsx")

    segmentos = mapear_contorno(dwg)
    if not segmentos:
        eag.msg.Print("Aviso: Nenhum segmento de contorno encontrado. Verifique a cor/nivel da borda da viga.")
        return

    ferros = extrair_ferros(dwg, eag)
    if not ferros:
        eag.msg.Print("Aviso: Nenhum ferro inteligente encontrado no desenho.")
        return

    classificar_ferros(ferros, segmentos)
    arquivo_salvo = gerar_relatorio(ferros, arquivo_excel)

    if arquivo_salvo == arquivo_excel:
        eag.msg.Print(f"Relatorio gerado com sucesso: {arquivo_salvo}")
    else:
        eag.msg.Print(
            f"ERRO: Tabela está aberta no Excel. Feche o arquivo '{arquivo_excel}' e tente novamente."
        )


if __name__ == "__main__":
    rodar_script()
