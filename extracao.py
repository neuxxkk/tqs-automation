import math
import os
from TQS import TQSDwg, TQSGeo, TQSEag, TQSJan
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# Fase 1: Preparação e Inicialização
# ==============================================================================

TOLERANCIA = 0.5  # tolerância em cm para comparação de cobrimento
CORES_BORDA_VIGA = {7, 1}  # cor branca (7) e vermelha (1) - faces e delimitadores da viga
ARQUIVO_EXCEL = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", "relatoriostqs", "relatorio_ferros.xlsx")
NOMEDWG = "desenho.DWG"


def inicializar_dwg(nomedwg):
    """Fase 1: Abre o desenho e configura a tolerância geométrica."""
    TQSGeo.SetPrecision(TOLERANCIA)

    dwg = TQSDwg.Dwg()
    if dwg.file.Open(nomedwg) != 0:
        print(f"Erro: Não foi possível abrir o desenho {nomedwg}")
        return None

    print(f"Desenho '{nomedwg}' aberto com sucesso.")
    return dwg


# ==============================================================================
# Fase 2: Mapeamento do Contorno da Viga
# ==============================================================================

def mapear_contorno(dwg):
    """Fase 2: Extrai todos os segmentos de reta da borda da viga."""
    segmentos = []

    dwg.iterator.Begin()
    while True:
        itipo = dwg.iterator.Next()

        if itipo == TQSDwg.DWGTYPE_EOF:
            break

        # Filtra apenas linhas e polilinhas
        if itipo not in (TQSDwg.DWGTYPE_LINE, TQSDwg.DWGTYPE_POLYLINE):
            continue

        # Filtra pelas cores de borda da viga (branca e vermelha/delimitadores)
        if dwg.iterator.color not in CORES_BORDA_VIGA:
            continue

        if itipo == TQSDwg.DWGTYPE_LINE:
            x1 = dwg.iterator.x1
            y1 = dwg.iterator.y1
            x2 = dwg.iterator.x2
            y2 = dwg.iterator.y2
            segmentos.append((x1, y1, x2, y2))

        elif itipo == TQSDwg.DWGTYPE_POLYLINE:
            npts = dwg.iterator.xySize
            if npts < 2:
                continue
            # Extrai pares de pontos consecutivos como segmentos
            x_ant, y_ant = dwg.iterator.GetPolylinePt(0)
            for ipt in range(1, npts):
                x_cur, y_cur = dwg.iterator.GetPolylinePt(ipt)
                segmentos.append((x_ant, y_ant, x_cur, y_cur))
                x_ant, y_ant = x_cur, y_cur

    print(f"Fase 2 concluída: {len(segmentos)} segmentos de contorno mapeados.")
    return segmentos


# ==============================================================================
# Fase 3: Leitura e Extração dos Ferros Inteligentes
# ==============================================================================

def extrair_ferros(dwg):
    """Fase 3: Lê todos os ferros inteligentes (IPOFER) e extrai geometria."""
    ferros = []

    dwg.iterator.Begin()
    while True:
        itipo = dwg.iterator.Next()

        if itipo == TQSDwg.DWGTYPE_EOF:
            break
        if itipo != TQSDwg.DWGTYPE_OBJECT:
            continue
        if dwg.iterator.objectName != "IPOFER":
            continue

        rebar = dwg.iterator.smartRebar
        cobrimento = rebar.cover

        # Extrair informações de tabela (posição, bitola, etc.)
        info_descr = []
        numdescr = rebar.RebarScheduleNumDescr()
        for idescr in range(numdescr):
            ipos, bitola, nfer, mult, itpcorbar, ivar, rdval, rddsc, compr, \
            igane, igand, observ, ilance, itipo99, icftppata, icorrido, \
            iluvai, iluvaf = rebar.RebarScheduleInfo(idescr)
            info_descr.append({
                "ipos": ipos,
                "bitola": bitola,
                "nfer": nfer,
                "compr": compr,
                "observ": observ,
            })

        # Extrair vértices de cada inserção do ferro
        pontos_insercao = []
        n_ins = rebar.GetInsertionNumber()
        for idx_ins in range(n_ins):
            resultado = rebar.GetInsertionPoints(idx_ins)
            # A API pode retornar uma tupla (npts, ...) em vez de um inteiro
            npts = resultado[0] if isinstance(resultado, tuple) else resultado
            for ipt in range(npts):
                pt = rebar.GetInsertionPoint(idx_ins, ipt)
                # A API pode retornar (x, y), (x, y, z) ou (status, x, y, ...)
                x, y = pt[0], pt[1]
                pontos_insercao.append((x, y))

        # Extrair informações de faixas de distribuição
        faixas = []
        numfaixas = rebar.GetRebarDistrNum()
        for ifaixa in range(numfaixas):
            icfes1, angfai, xpt1, ypt1, xpt2, ypt2, xcot, ycot, \
            ifdcotc, iflnfr, iflpos, iflbit, iflesp, \
            icentr, iquebr, ordem, k32vigas, k41vigas, \
            ilinexten, ilinchama, itpponta, espac, escxy = rebar.GetRebarDistrInfo(ifaixa)
            comp_faixa = math.hypot(xpt2 - xpt1, ypt2 - ypt1)
            faixas.append({
                "nferros": icfes1,
                "angulo": angfai,
                "comprimento": comp_faixa,
                "espacamento": espac,
            })

        # Verificar alternância pelo método correto da API
        alternado = rebar.alternatingMode == TQSDwg.ICPCAL
        fator_alternancia = rebar.alternatingFactor if alternado else None

        ferros.append({
            "rebar": rebar,
            "cobrimento": cobrimento,
            "info_descr": info_descr,
            "pontos": pontos_insercao,
            "faixas": faixas,
            "alternado": alternado,
            "fator_alternancia": fator_alternancia,
        })

    print(f"Fase 3 concluída: {len(ferros)} ferros inteligentes extraídos.")
    return ferros


# ==============================================================================
# Fase 4: Motor de Cruzamento Geométrico
# ==============================================================================

def distancia_ponto_segmento(xp, yp, x1, y1, x2, y2):
    """Calcula a menor distância de um ponto a um segmento finito de reta."""
    dx = x2 - x1
    dy = y2 - y1
    comp2 = dx * dx + dy * dy
    if comp2 == 0:
        # Segmento degenerado (ponto)
        return math.sqrt((xp - x1) ** 2 + (yp - y1) ** 2)
    # Parâmetro t da projeção no segmento [0,1]
    t = ((xp - x1) * dx + (yp - y1) * dy) / comp2
    t = max(0.0, min(1.0, t))
    # Ponto mais próximo no segmento
    px = x1 + t * dx
    py = y1 + t * dy
    return math.sqrt((xp - px) ** 2 + (yp - py) ** 2)


def ponto_na_borda(xp, yp, segmentos, cobrimento):
    """Verifica se um ponto do ferro está dentro da distância de cobrimento de algum segmento."""
    limite = cobrimento + TOLERANCIA
    for (x1l, y1l, x2l, y2l) in segmentos:
        dist = distancia_ponto_segmento(xp, yp, x1l, y1l, x2l, y2l)

        if dist < limite + 1e-6:
            return True

    return False


def classificar_ferros(ferros, segmentos):
    """Fase 4: Classifica cada ferro como 'borda', 'alternado' ou 'interno simples'."""
    for ferro in ferros:
        cobrimento = ferro["cobrimento"]
        ferro["eh_borda"] = any(
            ponto_na_borda(xp, yp, segmentos, cobrimento)
            for (xp, yp) in ferro["pontos"]
        )

    n_borda = sum(1 for f in ferros if f["eh_borda"])
    n_alt = sum(1 for f in ferros if not f["eh_borda"] and f["alternado"])
    n_simp = len(ferros) - n_borda - n_alt
    print(f"Fase 4 concluída: {n_borda} borda | {n_alt} alternados | {n_simp} internos simples.")


# ==============================================================================
# Fase 5: Relatório 
# ==============================================================================

def gerar_relatorio(ferros):
    """Fase 5: Gera planilha Excel com todos os ferros agrupados por posição."""
    ferros_ordenados = sorted(ferros, key=lambda f: f["info_descr"][0]["ipos"] if f["info_descr"] else 999)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ferros"

    # --- Estilos ---
    cinza     = PatternFill("solid", fgColor="D9D9D9")
    azul_claro= PatternFill("solid", fgColor="BDD7EE")
    verde     = PatternFill("solid", fgColor="C6EFCE")
    amarelo   = PatternFill("solid", fgColor="FFEB9C")
    laranja   = PatternFill("solid", fgColor="F4B183")
    vermelho  = PatternFill("solid", fgColor="FF9999")
    negrito   = Font(bold=True)
    centro    = Alignment(horizontal="center", vertical="center")
    borda_fina = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    def estilizar(cell, fill=None, bold=False, align=True):
        if fill:
            cell.fill = fill
        if bold:
            cell.font = Font(bold=True)
        if align:
            cell.alignment = centro
        cell.border = borda_fina

    # --- Cabeçalho ---
    cabecalho = ["Posição", "Quantidade", "Espaçamento (cm)", "Compr. Faixa (cm)", "Tipo"]
    ws.append(cabecalho)
    for col, _ in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col)
        estilizar(cell, fill=cinza, bold=True)

    # --- Linhas de dados ---
    cores_tipo = {"BORDA": vermelho, "ALTERNADO": amarelo, "INTERNO": azul_claro}

    soma_geral = soma_alt = soma_simp = soma_borda = 0.0
    pos_anterior = None

    for ferro in ferros_ordenados:
        if ferro["eh_borda"]:
            tipo = "BORDA"
        elif ferro["alternado"]:
            fator = ferro["fator_alternancia"]
            tipo = f"ALT(1/{fator})" if fator else "ALTERNADO"
        else:
            tipo = "INTERNO"

        nfer = ferro["info_descr"][0]["nfer"] if ferro["info_descr"] else 0
        comp_faixa_total = sum(f["comprimento"] for f in ferro["faixas"])
        espacamento = comp_faixa_total / (nfer - 1) if nfer > 1 else 0.0

        # Linha separadora entre grupos de posição
        ipos = ferro["info_descr"][0]["ipos"] if ferro["info_descr"] else None
        if pos_anterior is not None and ipos != pos_anterior:
            ws.append(["", "", "", "", ""])  # linha vazia separadora
        pos_anterior = ipos

        for info in ferro["info_descr"]:
            row = [
                f"N{info['ipos']}",
                info["nfer"],
                round(espacamento, 2),
                round(comp_faixa_total, 2),
                tipo,
            ]
            ws.append(row)
            fill = cores_tipo.get("BORDA" if ferro["eh_borda"] else
                                  ("ALTERNADO" if ferro["alternado"] else "INTERNO"), None)
            for col in range(1, 6):
                estilizar(ws.cell(ws.max_row, col), fill=fill)

        # Acumuladores de total
        soma_geral += comp_faixa_total
        if ferro["eh_borda"]:
            soma_borda += comp_faixa_total
        elif ferro["alternado"]:
            soma_alt += comp_faixa_total
        else:
            soma_simp += comp_faixa_total

    # --- Totais ---
    ws.append(["", "", "", "", ""])
    totais = [
        ("Total Geral (cm)",      soma_geral,  verde),
        ("Total Alternados (cm)", soma_alt,    amarelo),
        ("Total Simples (cm)",    soma_simp,   azul_claro),
        ("Total Borda (cm)",      soma_borda,  vermelho),
    ]
    for label, valor, fill in totais:
        ws.append([label, "", "", round(valor, 2), ""])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        for col in range(1, 6):
            estilizar(ws.cell(r, col), fill=fill, bold=True)

    # --- Larguras de coluna ---
    larguras = [14, 13, 18, 20, 14]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(ARQUIVO_EXCEL), exist_ok=True)
    wb.save(ARQUIVO_EXCEL)
    print(f"Planilha gerada: '{ARQUIVO_EXCEL}'")


# ==============================================================================
# Ponto de entrada principal
# ==============================================================================

# ==============================================================================
# Ponto de entrada via botão do TQS (EAG)
# ==============================================================================

def rodar_script(eag, tqsjan):
    """Função chamada pelo botão do TQS. Usa o desenho já aberto no editor."""
    eag.msg.Print("=== Identificação de Ferros de Borda ===")

    TQSGeo.SetPrecision(TOLERANCIA)

    # Usa o desenho já aberto na tela (desenho.DWG)
    dwg = tqsjan.dwg

    # Fase 2: Mapeamento do contorno da viga
    segmentos = mapear_contorno(dwg)
    if not segmentos:
        eag.msg.Print("Aviso: Nenhum segmento de contorno encontrado. "
                      "Verifique a cor/nível da borda da viga.")
        return

    # Fase 3: Extração dos ferros inteligentes
    ferros = extrair_ferros(dwg)
    if not ferros:
        eag.msg.Print("Aviso: Nenhum ferro inteligente encontrado no desenho.")
        return

    # Fase 4: Cruzamento geométrico
    classificar_ferros(ferros, segmentos)

    # Fase 5: Relatório
    gerar_relatorio(ferros)

    eag.msg.Print(f"Relatório gerado com sucesso: {ARQUIVO_EXCEL}")


# ==============================================================================
# Ponto de entrada standalone (fora do TQS)
# ==============================================================================

def main():
    print("=== Identificação de Ferros de Borda ===\n")

    # Fase 1: Inicialização
    dwg = inicializar_dwg(NOMEDWG)
    if dwg is None:
        return

    # Fase 2: Mapeamento do contorno da viga
    segmentos = mapear_contorno(dwg)
    if not segmentos:
        print("Aviso: Nenhum segmento de contorno encontrado. "
              "Verifique a cor/nível da borda da viga.")
        return

    # Fase 3: Extração dos ferros inteligentes
    ferros = extrair_ferros(dwg)
    if not ferros:
        print("Aviso: Nenhum ferro inteligente encontrado no desenho.")
        return

    # Fase 4: Cruzamento geométrico
    classificar_ferros(ferros, segmentos)

    # Fase 5: Relatório
    gerar_relatorio(ferros)


if __name__ == "__main__":
    main()